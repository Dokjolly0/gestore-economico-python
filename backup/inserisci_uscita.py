import datetime
import openpyxl
from openpyxl.styles import PatternFill
from function import *

def inserisci_uscita():
    try:
        tipologia = input("Inserisci una tipologia di uscita: ")

        importo = validate_input_float(0, 100000.0, "Inserisci l'importo dell'uscita: ")
        importo -= importo * 2
        giorno = get_day()
        data = get_date()
        data_format = data.strftime("%d/%m/%Y")
        descrizione = input("Inserisci una descrizione: ")
        try:
            # Cerca se file Excel esiste
            workbook = openpyxl.load_workbook('data_economy.xlsx')
            # Se esiste, apri il foglio di lavoro attivo
            sheet = workbook.active
        except:
            # Se non esiste, crea un nuovo file Excel
            workbook = openpyxl.Workbook()
            # Crea un nuovo foglio di lavoro
            sheet = workbook.active
            # Scrivi le intestazioni
            sheet.append(['Tipologia', 'Importo', 'Giorno', 'Data', 'Descrizione'])

        # Scrivi i dati in una nuova riga
        new_row = [tipologia, importo, giorno, data_format, descrizione]
        # Aggiungi la nuova riga all'intestazione corretta (posizione 0 Tipo, 1 Importo, 2 Giorno, 3 Data, 4 Descrizione)
        sheet.append(new_row)

        # Applica colore alla nuova riga (es. giallo chiaro)
        fill = PatternFill(start_color="ff0000", end_color="ff0000", fill_type="solid")
        for cell in sheet[sheet.max_row]:
            cell.fill = fill

        sheet = ridimensiona_file_excel(sheet)

        # Salva il file Excel
        workbook.save('data_economy.xlsx')
    except KeyboardInterrupt:
        print("\nUscita dal programma.")
        exit()
    except Exception as e:
        print("Errore: ", e)
