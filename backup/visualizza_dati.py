from function import *
import openpyxl
import tabulate
from openpyxl.styles import PatternFill

def visualizza_dati():
    try:
        choise_type = input("Scegli una categoria di movimenti da visualizzare o premi enter per visualizzare tutta la lisra movimenti: ")
        if choise_type == "":
            tipologia = None
        else:
            tipologia = choise_type
        ##########################################
        ########### filtra per mese ##############
        ##########################################
        try:
            # Cerca se file Excel esiste
            workbook = openpyxl.load_workbook('data_economy.xlsx')
            # Se esiste, apri il foglio di lavoro attivo
            sheet = workbook.active
        except:
            print("Il file Excel non esiste.")
            exit()
        valore_tipologia = []
        valore_movimenti = []
        valore_giorni = []
        valore_date = []
        valore_descrizione = []
        data = []
        n_righe = sheet.max_row
        n_colonne = sheet.max_column
        if n_righe == 0:
            print("Non ci sono dati da visualizzare.")
            exit()
        else:
            valore_tipologia = estrai_colonna(sheet, "A")
            valore_tipologia.remove('Tipologia')
            valore_movimenti = estrai_colonna(sheet, "B")
            valore_movimenti.remove('Importo')
            valore_giorni = estrai_colonna(sheet, "C")
            valore_giorni.remove('Giorno')
            valore_date = estrai_colonna(sheet, "D")
            valore_date.remove('Data')
            valore_descrizione = estrai_colonna(sheet, "E")
            valore_descrizione.remove('Descrizione')
            data = [valore_tipologia, valore_movimenti, valore_giorni, valore_date, valore_descrizione]

        # Stampa i dati in una tabella e crea un dizionario
        collection = []
        header = ["Tipologia", "Importo", "Giorno", "Data", "Descrizione"]
        index = 0
        while True:
            if tipologia != None:
                try:
                    if (data[0][index] != tipologia):
                        index += 1
                        continue
                except IndexError:
                    break

            try:
                #print(data[0][index])
                item = {}
                for row in data:
                    item[header[data.index(row)]] = row[index]
                collection.append(item)
                index += 1
            except IndexError:
                break
            except Exception as e:
                print("Errore: ", e)
                exit()

        print(tabulate.tabulate(collection, headers="keys", tablefmt="grid"))
        saldo_totale = round(sum(valore_movimenti), 2)
        if tipologia == None:
            print(f"\nSaldo totale: {saldo_totale} euro.")
        else:
            total = 0
            for item in collection:
                total += float(item["Importo"])
            print(f"\nTotale {tipologia}: {round(total, 2)} euro.")
            print(f"Saldo totale: {round(sum(valore_movimenti), 2)} euro.")
        return collection

    except KeyboardInterrupt:
        print("\nUscita dal programma.")
        exit()
    except Exception as e:
        print("Errore: ", e)
        exit()
