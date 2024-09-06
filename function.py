import datetime
import os
import sys
import openpyxl
from openpyxl.styles import PatternFill, Border, Side
import unicodedata

def exit():
    sys.exit()

def validate_input(start: int, end: int, text = "Scegli un'opzione: ", exclude_value = None):
    if isinstance(start, int) and isinstance(end, int):
        while True:
            choise_input = input(text)
            try:
                choise = int(choise_input)
                if start <= choise <= end:
                    return choise
            except ValueError:
                if choise_input == exclude_value:
                    if isinstance(exclude_value, list):
                        for element in exclude_value:
                            if element == choise_input:
                                return element
                    return choise_input
                else:
                    continue
            except KeyboardInterrupt:
                print("\nUscita dal programma.")
                exit()
            except:
                continue
            
def validate_input_float(start: float, end: float, text = "Scegli un'opzione: ", exclude_value = None):
    if (isinstance(start, int) or isinstance(start, float)) and (isinstance(end, int) or isinstance(end, float)):
        while True:
            choise_input = input(text)
            try:
                choise = float(choise_input)
                if start <= choise <= end:
                    return choise
            except ValueError:
                if choise_input == exclude_value:
                    if isinstance(exclude_value, list):
                        for element in exclude_value:
                            if element == choise_input:
                                return element
                    return choise_input
                else:
                    continue
            except KeyboardInterrupt:
                print("\nUscita dal programma.")
                exit()
            except:
                continue

# Funzione per rimuovere gli accenti
def remove_accents(input_str):
    return ''.join(c for c in unicodedata.normalize('NFD', input_str) if unicodedata.category(c) != 'Mn')

def get_date():
    data_input = input("Inserisci la data di accredito dell'importo o enter per la data attuale (gg/mm/aaaa): ").lower()
    if data_input == '':
        data = datetime.datetime.now()
        return data
    else:
        try:
            data = datetime.datetime.strptime(data_input, "%d/%m/%Y")
            return data
        except:
            print("Formato della data non corretto, reinseriscilo.")
            get_date()

def get_day_to_index(index: int):
    if isinstance(index, int) and 0 <= index <= 6:
        giorni = ["lunedì", "martedì", "mercoledì", "giovedì", "venerdì", "sabato", "domenica"]
        return giorni[index]

def get_day():
    giorni = ["lunedì", "martedì", "mercoledì", "giovedì", "venerdì", "sabato", "domenica"]
    giorno_input = input("Inserisci il giorno della settimana o enter per il giorno attuale: ").lower()

    if giorno_input == '':
        index = datetime.datetime.now().weekday()
        return get_day_to_index(index)

    # Rimuovi gli accenti dall'input
    giorno_input_normalized = remove_accents(giorno_input)

    # Confronta l'input normalizzato con i giorni normalizzati
    for giorno in giorni:
        if giorno_input_normalized == remove_accents(giorno):
            return giorno  # Restituisce il giorno corrispondente
    # Se nessun giorno corrisponde, mostra un errore
    print("Giorno non valido, riprova.")
    return get_day()  # Richiama la funzione per chiedere di nuovo l'input
        
def ridimensiona_file_excel(sheet):
    # Ridimensiona automaticamente la larghezza delle colonne
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter  # Ottieni la lettera della colonna (es. 'A')
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)  # Aggiungi un po' di spazio extra
        sheet.column_dimensions[column].width = adjusted_width
    return sheet


def estrai_colonna(sheet, index_colonna):
    # Inizializza una lista vuota per salvare i valori
    valori_colonna = []

    # Itera attraverso tutte le righe del foglio di lavoro per la colonna specificata
    for cell in sheet[index_colonna]:
        if cell.value is not None:  # Aggiungi solo valori non nulli
            valori_colonna.append(cell.value)
    return valori_colonna

def conta_righe_colonna(sheet, colonna_lettera):
    # Conta le celle non vuote nella colonna specificata
    conteggio = 0
    for cell in sheet[colonna_lettera]:
        if cell.value is not None:
            conteggio += 1
    return conteggio

def filter_date(date_list, mese: int, anno: int):
    try:
        # Verifica che mese e anno siano interi
        if (not isinstance(mese, int)) or (not isinstance(anno, int)):
            print(f"Mese: {mese} \nAnno: {anno}")
            print("Mese o anno non corretti.")
            return False

        # Converte le stringhe in oggetti datetime, se necessario
        date_objects = []
        for data in date_list:
            if isinstance(data, str):
                # Prova a convertire la stringa in un oggetto datetime con formato 'DD/MM/YYYY'
                try:
                    data = datetime.datetime.strptime(data, "%d/%m/%Y")  # Formato 'DD/MM/YYYY'
                except ValueError:
                    print(f"Formato data non valido: {data}")
                    continue
            date_objects.append(data)

        # print(f"Oggetti datetime convertiti: {date_objects}")
        # print(f"Filtriamo per mese: {mese}, anno: {anno}")

        # Filtra le date in base al mese e anno
        filtered_dates = [data for data in date_objects if data.month == mese and data.year == anno]

        #print(f"Date filtrate: {filtered_dates}")

        # Riconverte le date filtrate in stringhe formato 'DD/MM/YYYY'
        return [data.strftime("%d/%m/%Y") for data in filtered_dates]

    except KeyboardInterrupt:
        print("\nUscita dal programma.")
        exit()
    except Exception as e:
        print(f"Errore filtra date -> {str(e)}")

def salva_excel_visualizzato(data: []):
    if isinstance(data, list):
        try:
            # Nuovo file excel
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.append(['Tipologia', 'Importo', 'Giorno', 'Data', 'Descrizione'])

            black_border = Border(
                left=Side(border_style="thin", color="000000"),
                right=Side(border_style="thin", color="000000"),
                top=Side(border_style="thin", color="000000"),
                bottom=Side(border_style="thin", color="000000")
            )

            for item in data:
                #Valori da inserire
                tipologia = item["Tipologia"]
                importo = item["Importo"]
                giorno = item["Giorno"]
                data = item["Data"]
                descrizione = item["Descrizione"]

                # Scrivi i dati in una nuova riga
                new_row = [tipologia, importo, giorno, data, descrizione]
                sheet.append(new_row)

                # Colora le caselle e ridimensiona il file Excel
                fill = PatternFill(start_color="008f39", end_color="008f39", fill_type="solid")
                for cell in sheet[sheet.max_row]:
                    cell.fill = fill
                    cell.border = black_border
                sheet = ridimensiona_file_excel(sheet)

            # Salva il file Excel
            nome = input("Inserisci il nome del file: ")
            print("Seleziona il percorso dove salvare il file Excel.")
            print("1- Salva in C:\\MyDatabase")
            print("2- Salva in nella posizione attuale del file")
            print("3- Salva in un percorso personalizzato")
            choise = validate_input(1, 3, "Seleziona un opzione: ")
            if choise == 1:
                full_path = os.path.join(r"C:\MyDatabase", nome + ".xlsx")
                workbook.save(full_path)
            elif choise == 2:
                file_path = os.path.abspath(__file__)
                directory_path = os.path.dirname(file_path)
                full_path = os.path.join(directory_path, nome + ".xlsx")
                workbook.save(full_path)
            elif choise == 3:
                path = input("Inserisci il percorso: ")
                while not os.path.exists(path) or not os.path.isdir(path):
                    print("Percorso non valido.")
                    path = input("Inserisci il percorso: ")
                    if os.path.exists(path) and os.path.isdir(path):
                        break
                full_path = os.path.join(path, nome + ".xlsx")
                workbook.save(full_path)
        except KeyboardInterrupt:
            print("\nUscita dal programma.")
            exit()
        except Exception as e:
            print(f"Errore salvataggio Excel -> {str(e)}")

#salva_excel_visualizzato([{'Tipologia': 'Entrata stipendio pizzeria', 'Importo': 321, 'Giorno': 'giovedì', 'Data': '05/08/2020', 'Descrizione': 'Mese e anno diversi'}, {'Tipologia': 'Entrata stipendio pizzeria', 'Importo': 123, 'Giorno': 'giovedì', 'Data': '05/08/2024', 'Descrizione': 'Mese diverso anno uguale'}, {'Tipologia': 'Entrata stipendio pizzeria', 'Importo': 12, 'Giorno': 'giovedì', 'Data': '05/09/2024', 'Descrizione': 'Anno e mese uguali'}])
# date_list = ["05/08/2020", "12/09/2021", "05/08/2020", "01/01/2022", "05/09/2024"]
# print(filter_date(date_list, 9, 2024))  # Filtra per agosto 2020