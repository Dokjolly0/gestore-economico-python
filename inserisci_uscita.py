import datetime
import os

import openpyxl
from openpyxl.styles import PatternFill
from function import *

def inserisci_uscita(path=r"C:\MyDatabase"):
    try:
        if not os.path.exists(path):
            print("Path non esistente.")
            exit()
            
        print("Inserisci una tipologia di uscita tra le seguenti: ")
        print("1- Abbonamento online.")
        print("2- Prelevamento atm.")
        print("3- Giochi.")
        print("4- Pagamento amazon.")
        print("5- Altra tipologia di uscita.")
        tipologia = validate_input(1, 5, "Inserisci una tipologia di uscita: ")
        if tipologia == 1:
            tipologia = 'Abbonamento online'
        elif tipologia == 2:
            tipologia = 'Prelevamento atm'
        elif tipologia == 3:
            tipologia = 'Giochi'
        elif tipologia == 4:
            tipologia = 'Pagamento amazon'
        elif tipologia == 5:
            tipologia = input("Inserisci la tipologia di uscita: ")

        importo = validate_input_float(0, 100000.0, "Inserisci l'importo dell'uscita: ")
        importo -= importo * 2
        giorno = get_day()
        data = get_date()
        data_format = data.strftime("%d/%m/%Y")
        descrizione = input("Inserisci una descrizione: ")
        try:
            # Cerca se file Excel esiste
            full_path = os.path.join(path, 'data_economy.xlsx')
            workbook = openpyxl.load_workbook(full_path)
            # Se esiste, apri il foglio di lavoro attivo
            sheet = workbook.active
        except:
            # Se non esiste, crea un nuovo file Excel
            workbook = openpyxl.Workbook()
            # Crea un nuovo foglio di lavoro
            sheet = workbook.active
            # Scrivi le intestazioni
            sheet.append(['Tipologia', 'Importo', 'Giorno', 'Data', 'Descrizione'])

        black_border = Border(
            left=Side(border_style="thin", color="000000"),
            right=Side(border_style="thin", color="000000"),
            top=Side(border_style="thin", color="000000"),
            bottom=Side(border_style="thin", color="000000")
        )

        # Scrivi i dati in una nuova riga
        new_row = [tipologia, importo, giorno, data_format, descrizione]
        # Aggiungi la nuova riga all'intestazione corretta (posizione 0 Tipo, 1 Importo, 2 Giorno, 3 Data, 4 Descrizione)
        sheet.append(new_row)

        # Applica colore alla nuova riga (es. giallo chiaro)
        fill = PatternFill(start_color="ff0000", end_color="ff0000", fill_type="solid")
        for cell in sheet[sheet.max_row]:
            cell.fill = fill
            cell.border = black_border
        sheet = ridimensiona_file_excel(sheet)

        # Salva il file Excel
        full_path = os.path.join(path, 'data_economy.xlsx')
        workbook.save(full_path)
    except KeyboardInterrupt:
        print("\nUscita dal programma.")
        exit()
    except Exception as e:
        print("Errore: ", e)
