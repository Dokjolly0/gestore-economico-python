import os
import datetime
from function import *
import openpyxl
import tabulate
from openpyxl.styles import PatternFill

def visualizza_dati(path=r"C:\MyDatabase"):
    try:
        if not os.path.exists(path):
            print("Path non esistente.")
            exit()
        choise_type = input("Scegli una categoria di movimenti da visualizzare o premi enter per visualizzare tutta la lisra movimenti: ")
        if choise_type == "":
            tipologia = None
        else:
            tipologia = choise_type

        filter_date_bool = False
        choise_date = input("Vuoi filtrare i movimenti per data? (y/n) ").lower()
        if choise_date == "y" or choise_date == "s":
            filter_date_bool = True
            mese = validate_input(1, 12, "Inserisci il mese o enter per il mese corrente: ", '')
            anno = validate_input(1900, 2100, "Inserisci l'anno o enter per l'anno corrente: ", '')
            if mese == "":
                mese = datetime.datetime.now().month
            if anno == "":
                anno = datetime.datetime.now().year
        else:
            mese = None
            anno = None

        try:
            # Cerca se file Excel esiste
            full_path = os.path.join(path, 'data_economy.xlsx')
            workbook = openpyxl.load_workbook(full_path)
            # Se esiste, apri il foglio di lavoro attivo
            sheet = workbook.active
        except:
            print("Il file Excel non esiste.")
            exit()
        valore_tipologia = []
        valore_movimenti = []
        valore_giorni = []
        valore_date = []
        valore_descrizione = []
        data = []
        n_righe = sheet.max_row
        n_colonne = sheet.max_column
        if n_righe == 0:
            print("Non ci sono dati da visualizzare.")
            exit()
        else:
            valore_tipologia = estrai_colonna(sheet, "A")
            valore_tipologia.remove('Tipologia')
            valore_movimenti = estrai_colonna(sheet, "B")
            valore_movimenti.remove('Importo')
            valore_giorni = estrai_colonna(sheet, "C")
            valore_giorni.remove('Giorno')
            valore_date = estrai_colonna(sheet, "D")
            valore_date.remove('Data')
            valore_descrizione = estrai_colonna(sheet, "E")
            valore_descrizione.remove('Descrizione')
            data = [valore_tipologia, valore_movimenti, valore_giorni, valore_date, valore_descrizione]

        # Stampa i dati in una tabella e crea un dizionario
        collection = []
        header = ["Tipologia", "Importo", "Giorno", "Data", "Descrizione"]
        index = 0
        while True:
            if tipologia != None:
                try:
                    if (data[0][index] != tipologia):
                        index += 1
                        continue
                except IndexError:
                    break

            if (mese != None) or (anno != None):
                filter_valore_date = filter_date(valore_date, mese, anno)
                valore_date = filter_valore_date

            try:
                #print(data[0][index])
                item = {}
                for row in data:
                    if data[3][index] in valore_date:
                        item[header[data.index(row)]] = row[index]
                if item:  # Check if item has any data
                    collection.append(item)
                #collection.append(item)
                index += 1
            except IndexError:
                break
            except Exception as e:
                print("Erroreee: ", e)
                exit()

        print(tabulate.tabulate(collection, headers="keys", tablefmt="grid"))
        saldo_totale = round(sum(valore_movimenti), 2)

        total = 0
        for item in collection:
            total += float(item["Importo"])
        if round(total, 2) != round(sum(valore_movimenti), 2):
            print(f"\nTotale parziale: {round(total, 2)} euro.")
        print(f"Saldo totale: {round(sum(valore_movimenti), 2)} euro.")
        return collection

    except KeyboardInterrupt:
        print("\nUscita dal programma.")
        exit()
    except Exception as e:
        print("Errore: ", e)
        exit()
