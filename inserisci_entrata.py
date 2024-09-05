import datetime
import os
import openpyxl
from openpyxl.styles import PatternFill
from function import *

def inserisci_entrata(path=r"C:\MyDatabase"):
    try:
        if not os.path.exists(path):
            print("Path non esistente.")
            exit()
        print("Inserisci una tipologia di entrata tra le seguenti: ")
        print("1- Entrata stipendio pizzeria.")
        print("2- Entrata mance pizzeria.")
        print("3- Entrata stipendio lavoro.")
        print("4- Altra tipologia di entrata.")
        choise = validate_input(1, 4)

        tipologia = ''
        if choise == 1:
            tipologia = 'Entrata stipendio pizzeria'
        elif choise == 2:
            tipologia = 'Entrata mance pizzeria'
        elif choise == 3:
            tipologia = 'Entrata stipendio lavoro'
        elif choise == 4:
            tipologia = input("Inserisci la tipologia di entrata: ")

        importo = validate_input_float(0, 10000.0, "Inserisci l'importo dell'entrata: ")
        giorno = get_day()
        data = get_date()
        data_format = data.strftime("%d/%m/%Y")
        descrizione = input("Inserisci una descrizione: ")
        try:
            # Cerca se file Excel esiste
            full_path = os.path.join(path, 'data_economy.xlsx')
            workbook = openpyxl.load_workbook(full_path)
            # Se esiste, apri il foglio di lavoro attivo
            sheet = workbook.active
        except:
            # Se non esiste, crea un nuovo file Excel
            workbook = openpyxl.Workbook()
            # Crea un nuovo foglio di lavoro
            sheet = workbook.active
            # Scrivi le intestazioni
            sheet.append(['Tipologia', 'Importo', 'Giorno', 'Data', 'Descrizione'])

        # Scrivi i dati in una nuova riga
        new_row = [tipologia, importo, giorno, data_format, descrizione]
        # Aggiungi la nuova riga all'intestazione corretta (posizione 0 Tipo, 1 Importo, 2 Giorno, 3 Data, 4 Descrizione)
        #print(f"Sheet prima: {sheet.max_row}")
        sheet.append(new_row)
        #print(f"Sheet dopo: {sheet.max_row}")

        # Applica colore alla nuova riga (es. giallo chiaro)
        fill = PatternFill(start_color="008f39", end_color="008f39", fill_type="solid")
        for cell in sheet[sheet.max_row]:
            cell.fill = fill

        sheet = ridimensiona_file_excel(sheet)

        # Salva il file Excel
        full_path = os.path.join(path, 'data_economy.xlsx')
        workbook.save(full_path)
    except KeyboardInterrupt:
        print("\nUscita dal programma.")
        exit()
    except Exception as e:
        print("Errore: ", e)
